from openpyxl import Workbook
import re

EXCEPTION_LIST = [' ', 'AEP-84 Volume II', 'Edition A Version 1']
re_page_number = re.compile(r'([0-9]+ - [0-9]+)')    # 페이지 번호 여부 확인
re_section_number = re.compile(r'(([0-9+]+)(\.[0-9]+)?[^\.])')  # 장절 번호 추출

def write_excel(filepath, text_from_pdf):
    wb = Workbook()
    ws = wb.active

    start_row = 1
    section_col = 1
    text_col = 2
    buffer = ''

    for page_number, text in enumerate(text_from_pdf):
        total_line = len(text.split('\n'))
        print('total_line: ', total_line)
        for line_number, line in enumerate(text.split('\n')):
            line_strip = line.strip()
            print('line_strip: ', line_strip)
            if (not line_strip in EXCEPTION_LIST) and (not line_strip == ''):                
                ascii = ord(line_strip[0])                
                print('ascii: ', ascii)
                print('type ascii: ', type(ascii))
                if (48 <= ascii) and (ascii <= 57):     # 숫자로 시작
                    # 숫자로 시작하지 않은 문자가 buffer에 쌓여 있는 경우 이를 현재 행에 출력
                    if buffer != '':
                        ws.cell(row=start_row, column=text_col, value=buffer)
                        start_row += 1
                        buffer = ''                    
                    try:
                        # 쪽 번호인 경우 skip
                        is_page_number = re_page_number.match(line_strip)
                        if is_page_number:
                            if text.split('\n')[line_number + 1].strip() == 'Edition A Version 1':
                                continue
                    except:
                        pass
                        
                    ws.cell(row=start_row, column=text_col, value=line_strip)
                    start_row += 1

                else:
                    if buffer != '':
                        if ascii == 61623:  # 가운데 점
                            buffer = buffer + '\n· ' + line_strip[3:]
                        elif ascii == 45:
                            buffer = buffer + '\n' + line_strip
                        else:
                            buffer = buffer + ' ' + line_strip
                    else:
                        if ascii == 61623:  # 가운데 점
                            buffer = '· ' + line_strip[3:]
                        else:
                            buffer = line_strip

                    # 마지막 line일 경우
                    if line_number + 1 == total_line:
                        ws.cell(row=start_row, column=text_col, value=buffer)
                        start_row += 1
                        buffer = ''
                    continue

    wb.save(filename=filepath)